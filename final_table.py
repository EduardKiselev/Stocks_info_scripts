import sqlite3
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Список выбранных тикеров
selected_tickers = ['IMOEX',
                    'RGBI',
                    'X5',
                    'SBER',
                    'T',
                    '9618.HK',
                    '3798.HK',
                    '9988.HK',
                    '9888.HK',
                    '0219.HK',
                    'UGLD',
                    '2127.HK',
                    '2007.HK',
                    'NBIS',
                    'TMF',
                    'TLT',
                    'CNYRUB_TOM',
                    'USDRUBF',
                    '^GSPC',
                    '^IXIC',
                    '^DJI']

with open('cherry.txt','r') as cherry:
    cherries = [line.strip().split() for line in cherry.readlines()]

    cherry_list=[]
    for cherry in cherries:
        if cherry[1] == 'RU':
            name = cherry[2] + cherry[0]
            cherry_list.append(name)
        elif cherry[1] == 'US':
            name = cherry[2]
            cherry_list.append(name)
selected_tickers += cherry_list


print(selected_tickers)
# Подключение к БД
conn = sqlite3.connect("stocks.db")

# Чтение данных из таблицы `daily_analysis` с JOIN таблицы `countries`
query = """
SELECT da.*, c.country
FROM daily_analysis da
LEFT JOIN countries c ON da.ticker = c.ticker
"""
data = pd.read_sql(query, conn)

query = """
SELECT *
FROM indices_analysis
"""
indices_data = pd.read_sql(query, conn)
conn.close()


# Преобразуем столбец `date` в формат datetime
data["date"] = pd.to_datetime(data["date"])
indices_data["date"] = pd.to_datetime(indices_data["date"])

# Проверяем, передан ли аргумент даты
if len(sys.argv) > 1:
    date_arg = sys.argv[1]  # Получаем дату из аргументов командной строки
    try:
        # Преобразуем строку в объект datetime
        selected_date = pd.to_datetime(date_arg, format="%d-%m-%y")
    except ValueError:
        print("Неверный формат даты. Используйте формат DD-MM-YY.")
        sys.exit(1)
else:
    selected_date = datetime.today().strftime("%Y-%m-%d")

# Фильтруем данные по выбранной дате
filtered_data = data[data["date"] == selected_date]
filtered_indices = indices_data[indices_data["date"] == selected_date]

# Объединяем данные из `filtered_data` и `filtered_indices`
# Используем `ticker` как ключ для объединения

# Удаляем ненужные столбцы из `data`
columns_to_drop = ['open', 'high', 'low', 'volume', 'ISA_9', 'ISB_26', 'ITS_9', 'IKS_26','id','obv','obv_sma_21']
data_cleaned = filtered_data.drop(columns=columns_to_drop)

# Добавляем недостающие столбцы в `filtered_indices`
filtered_indices['cloud_flag'] = 0  # Добавляем столбец cloud_flag со значением 0
filtered_indices['obv_flag'] = 0    # Добавляем столбец obv_flag со значением 0
filtered_indices['country'] = 'ru_index'  # Добавляем столбец country со значением 'ru_index'

# Приводим столбцы filtered_indices к тому же порядку, что и в data_cleaned
filtered_indices = filtered_indices[data_cleaned.columns]



# Объединяем таблицы с помощью pd.concat
merged_data = pd.concat([data_cleaned, filtered_indices], ignore_index=True)

# Упорядочиваем данные по странам
order = ['Index', 'RU', 'US', 'EU', 'HongKong','ru_index']
merged_data["country"] = pd.Categorical(merged_data["country"], categories=order, ordered=True)
merged_data = merged_data.sort_values(by="country")



# Создаем таблицу с флагами
flags_table = merged_data[["ticker",
                           "country",
                           "close",
                           "ema_7",
                           "rsi_14",
                           "rsi_flag",
                           "sma_flag",
                           "macd_flag",
                           "obv_flag",
                           "ema_flag",
                           "cloud_flag"]]

# Добавляем столбец с суммой флагов
flags_table["total_flags"] = (
    flags_table["sma_flag"] +
    flags_table["macd_flag"] +
    flags_table["obv_flag"] +
    flags_table["rsi_flag"] +
    flags_table["ema_flag"] +
    flags_table["cloud_flag"]
)

# Округляем значения в столбцах `ema_7` и `rsi_14` до 2 знаков после запятой
flags_table["ema_7"] = flags_table["ema_7"].round(2)
flags_table["rsi_14"] = flags_table["rsi_14"].round(0)

# Создаем Excel-файл
wb = Workbook()
ws = wb.active

# Записываем данные в Excel
for r in dataframe_to_rows(flags_table, index=False, header=True):
    ws.append(r)

# Определяем цвета для раскрашивания
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Зеленый
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Красный

# Раскрашиваем строки в зависимости от значения `total_flags`
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    total_flags = row[11].value  # Столбец `total_flags` индексация с 0
    if total_flags is not None:
        if total_flags > 2:
            for cell in row:
                cell.fill = green_fill  # Зеленый для total_flags > 2
        elif total_flags < -2:
            for cell in row:
                cell.fill = red_fill  # Красный для total_flags < -2

# Группируем данные по стране и считаем сумму total_flags для каждой страны
country_flags_sum = flags_table.groupby("country")["total_flags"].sum().reset_index()
country_sum = flags_table.groupby("country")["total_flags"].count().reset_index()

# Переименуем столбцы для удобства
country_flags_sum.columns = ["country", "total_flags_sum"]
country_sum.columns = ["country", "total_flags_count"]

# Объединяем данные
summary_table = pd.merge(country_flags_sum, country_sum, on="country")

# Добавляем столбец с отношением total_flags_sum / total_flags_count
summary_table["flags_ratio"] = (summary_table["total_flags_sum"] / summary_table["total_flags_count"]).round(1)

# Создаем новый лист в Excel-файле для таблицы с суммой флагов по странам
ws_summary = wb.create_sheet(title="Country Flags Summary")

# Записываем данные в новый лист
for r in dataframe_to_rows(summary_table, index=False, header=True):
    ws_summary.append(r)

# Фильтруем данные по выбранным тикерам
selected_tickers_data = flags_table[flags_table['ticker'].isin(selected_tickers)]

print("SELECTED_TICKERS: " + selected_tickers)

ws_selected_tickers = wb.create_sheet(title="Selected Tickers")

for r in dataframe_to_rows(selected_tickers_data, index=False, header=True):
    ws_selected_tickers.append(r)

# Раскрашиваем строки в зависимости от значения `total_flags`
for row in ws_selected_tickers.iter_rows(min_row=2, max_row=ws_selected_tickers.max_row, min_col=1, max_col=ws_selected_tickers.max_column):
    total_flags = row[11].value  # Столбец `total_flags` индексация с 0
    if total_flags is not None:
        if total_flags > 2:
            for cell in row:
                cell.fill = green_fill  # Зеленый для total_flags > 2
        elif total_flags < -2:
            for cell in row:
                cell.fill = red_fill  # Красный для total_flags < -2

output_file = "Stocks_analisys/final_table_"+str(selected_date.strftime("%Y-%m-%d"))+".xlsx"
wb.save(output_file)
print(f"Таблица сохранена в файл {output_file}")
